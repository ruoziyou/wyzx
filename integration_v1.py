import pymysql
import pandas as pd
import openpyxl
import os

# 获取数据库名称和MySQL密码
db_name = input("数据库名称: ")
db_pw = input("MySQL密码: ")

# 连接MySQL并获取游标
conn = pymysql.connect(host="127.0.0.1", user="root", passwd=db_pw, db=db_name, charset="utf8")
cursor = conn.cursor()

# 获取输出表格的类型，检查是否合理并报错
export_type = input("输出表格类型（输入1或2）：")
if not (export_type == "1" or export_type == "2"):
    raise ValueError("表格类型错误")

# 获取输出表格的名称，检查是否合理并报错
export_name = input("输出文件名称：")
if len(export_name) < 6 or not export_name[-5:] == ".xlsx":
    raise ValueError("文件名称错误")

# 获取输出表格的路径，检查是否合理并报错
export_path = input("输出文件路径（如果没有路径要求不用输入）：")
if export_path == "":
    export_file = export_name
else:
    if not os.path.exists(export_path) or os.path.isfile(export_path):
        raise ValueError("文件路径错误")
    else:
        export_file = f"{export_path}/{export_name}"

# 第1种表格类型
if export_type == "1":
    # 获取要查询的姓名及对应id，并检查该id是否在数据库中
    input_uuid = input("登记人姓名：")
    cursor.execute(f"SELECT id FROM user WHERE uuid= '{input_uuid}'")
    input_id = cursor.fetchone()
    if input_id is None:
        raise ValueError("姓名错误")
    else:
        input_id = input_id[0]

    # 获取内网数据，按日期排序，并整合成DataFrame
    cursor.execute(f"SELECT create_time, ip, ggroup, ports FROM private WHERE user_id = {input_id} ORDER BY create_time")
    df_private = pd.DataFrame(cursor.fetchall(), columns=["启用日期", "IP", "项目组", "开放端口数量"])

    # 获取外网数据，按日期排序，并整合成DataFrame
    cursor.execute(f"SELECT create_time, ip, servers, ports FROM public WHERE user_id = {input_id} ORDER BY create_time")
    df_public = pd.DataFrame(cursor.fetchall(), columns=["启用日期", "IP", "项目组", "开放端口数量"])

    # 创建Excel文件，合并header的单元格，并输入表格的header
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("sheet1", 0)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
    sheet.cell(1, 1).value = f"{input_uuid}: {df_private.shape[0] + df_public.shape[0]}"  # 姓名：资产数量
    sheet.cell(2, 1).value = "内网资产"
    sheet.cell(2, 5).value = "外网资产"
    wb.save(export_file)

    # 输出内网数据和外网数据到Excel表格里面
    with pd.ExcelWriter(export_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        df_private.to_excel(writer, sheet_name="sheet1", index=False, header=True, startrow=2)
        df_public.to_excel(writer, sheet_name="sheet1", index=False, header=True, startrow=2, startcol=4)

# 第二种表格类型
else:
    # 创建Excel文件，合并header的单元格，并输入表格的header
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("sheet1", 0)
    sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
    sheet.cell(1, 3).value = "内网"
    sheet.cell(1, 6).value = "外网"
    wb.save(export_file)

    # 设置计数器（计数器设置为0是为了只在第一次输出DataFrame的时候输出header）
    counter = 0

    # 获取所有登记人和其日期
    cursor.execute("SELECT create_time, id, uuid FROM user ORDER BY create_time")
    user_info = cursor.fetchall()

    # 遍历全部登记人
    for user in user_info:
        # 获取该登记人的内网数据，按IP排序，并整合成DataFrame
        cursor.execute(f"SELECT ip, ggroup, ports FROM private WHERE user_id = {user[1]} ORDER BY ip")
        df_private = pd.DataFrame(cursor.fetchall(), columns=["IP", "项目组", "开放端口数量"])

        # 获取该登记人的外网数据，按IP排序，并整合成DataFrame
        cursor.execute(f"SELECT ip, servers, ports FROM public WHERE user_id = {user[1]} ORDER BY ip")
        df_public = pd.DataFrame(cursor.fetchall(), columns=["IP", "项目组", "开放端口数量"])

        # 计算当前登记人的内网/外网资源占据的行数
        lines = max(df_private.shape[0], df_public.shape[0])

        # 获取登记人信息并整合成对应行数的DataFrame
        df_user = pd.DataFrame(((user[0], user[2]),) * lines, columns=["日期", "登记人"])

        # 输出登记人信息，内网数据，和外网数据到Excel表格里面
        with pd.ExcelWriter(export_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            df_user.to_excel(writer, sheet_name="sheet1", index=False, header=not counter, startrow=counter + 1)
            df_private.to_excel(writer, sheet_name="sheet1", index=False, header=not counter, startrow=counter + 1, startcol=2)
            df_public.to_excel(writer, sheet_name="sheet1", index=False, header=not counter, startrow=counter + 1, startcol=5)

        # 更新计数器
        if counter == 0:
            counter += 1  # 由于第一次输出要打印header，避免后面输出会覆盖第一次的最后一行数据
        counter += lines

# 关闭游标和连接
cursor.close()
conn.close()
